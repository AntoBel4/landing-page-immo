#!/usr/bin/env python3
"""
===============================================================================
KIT MANDATAIRE 2026 — GÉNÉRATEUR DE FICHIERS EXCEL
===============================================================================
Design : Tech Premium 2026
Couleur Header : Bleu Nuit #0A0E27
Accent : Bleu #3B82F6

Génère 3 fichiers :
1. Tracker_Consentements_RGPD.xlsx — Suivi opt-in conforme loi 2026
2. Calculateur_Pret_Relais.xlsx — Coût du retard vendeur
3. Suivi_DPE_Coefficient.xlsx — Impact réforme 2.3 → 1.9
===============================================================================
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

# =============================================================================
# CONSTANTES DE STYLE — TECH PREMIUM 2026
# =============================================================================
BLEU_NUIT = "0A0E27"
ACCENT_BLEU = "3B82F6"
GRIS_CLAIR = "F8FAFC"
GRIS_BORDURE = "E2E8F0"
VERT_SUCCES = "10B981"
ROUGE_ALERTE = "EF4444"
ORANGE_WARNING = "F59E0B"
BLANC = "FFFFFF"

# Fonts
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color=BLANC)
FONT_BODY = Font(name='Calibri', size=10, color=BLEU_NUIT)
FONT_TITLE = Font(name='Calibri', size=16, bold=True, color=BLEU_NUIT)
FONT_SUBTITLE = Font(name='Calibri', size=10, italic=True, color="64748B")
FONT_ACCENT = Font(name='Calibri', size=11, bold=True, color=ACCENT_BLEU)
FONT_RESULT = Font(name='Calibri', size=12, bold=True, color=VERT_SUCCES)
FONT_ALERT = Font(name='Calibri', size=12, bold=True, color=ROUGE_ALERTE)

# Fills
FILL_HEADER = PatternFill(start_color=BLEU_NUIT, end_color=BLEU_NUIT, fill_type='solid')
FILL_ACCENT = PatternFill(start_color=ACCENT_BLEU, end_color=ACCENT_BLEU, fill_type='solid')
FILL_GRIS = PatternFill(start_color=GRIS_CLAIR, end_color=GRIS_CLAIR, fill_type='solid')
FILL_INPUT = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type='solid')
FILL_RESULT = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type='solid')
FILL_ALERT = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type='solid')
FILL_WARNING = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type='solid')

# Borders
BORDER_THIN = Border(
    left=Side(style='thin', color=GRIS_BORDURE),
    right=Side(style='thin', color=GRIS_BORDURE),
    top=Side(style='thin', color=GRIS_BORDURE),
    bottom=Side(style='thin', color=GRIS_BORDURE)
)
BORDER_ACCENT = Border(
    left=Side(style='medium', color=ACCENT_BLEU),
    right=Side(style='medium', color=ACCENT_BLEU),
    top=Side(style='medium', color=ACCENT_BLEU),
    bottom=Side(style='medium', color=ACCENT_BLEU)
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')


def style_header(cell):
    """Style header : fond bleu nuit, texte blanc"""
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.border = BORDER_THIN
    cell.alignment = ALIGN_CENTER


def style_body(cell, alternate=False):
    """Style body standard"""
    cell.font = FONT_BODY
    cell.border = BORDER_THIN
    cell.alignment = ALIGN_LEFT
    if alternate:
        cell.fill = FILL_GRIS


def style_input(cell):
    """Style cellule de saisie (fond bleu clair)"""
    cell.font = FONT_ACCENT
    cell.fill = FILL_INPUT
    cell.border = BORDER_ACCENT
    cell.alignment = ALIGN_CENTER


def adjust_columns(ws, widths_dict):
    """Ajuster largeur des colonnes"""
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width


# =============================================================================
# 1. TRACKER CONSENTEMENTS RGPD
# =============================================================================
def create_tracker_consentements(filepath):
    """
    Tracker de consentements conforme RGPD + Loi Opt-in 2026
    Colonnes : ID, Date Contact, Plateforme, Lien Annonce, Nom/Pseudo,
               Téléphone, Date Opt-in, Expiration +3ans, Preuve O/N, Statut, Notes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker Consentements"
    
    # === TITRE ===
    ws.merge_cells('A1:K1')
    ws['A1'] = "📋 TRACKER CONSENTEMENTS — LOI OPT-IN 11 AOÛT 2026"
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 35
    
    # === SOUS-TITRE ===
    ws.merge_cells('A2:K2')
    ws['A2'] = "Obligation légale : consentement préalable, vérifiable et documenté. Conservation : 3 ans minimum (RGPD Art. 7)"
    ws['A2'].font = FONT_SUBTITLE
    ws['A2'].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 20
    
    # === HEADERS (ligne 4) ===
    headers = [
        ("A4", "ID", 6),
        ("B4", "Date Contact", 14),
        ("C4", "Plateforme", 14),
        ("D4", "Lien Annonce", 35),
        ("E4", "Nom / Pseudo", 18),
        ("F4", "Téléphone", 14),
        ("G4", "Date Opt-in", 14),
        ("H4", "Expiration\n(+3 ans)", 14),
        ("I4", "Preuve\nStockée", 10),
        ("J4", "Statut", 16),
        ("K4", "Notes", 25),
    ]
    
    for cell_ref, title, width in headers:
        ws[cell_ref] = title
        style_header(ws[cell_ref])
        col_letter = cell_ref[0] if len(cell_ref) == 2 else cell_ref[:2]
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[4].height = 35
    
    # === EXEMPLES (lignes 5-7) ===
    examples = [
        (1, "15/01/2026", "Leboncoin", "https://www.leboncoin.fr/...", "Propriétaire_A", "06 XX XX XX XX", "16/01/2026", "", "✅ Oui", "RDV Planifié", "Rappeler mardi 14h"),
        (2, "18/01/2026", "SeLoger", "https://www.seloger.com/...", "Vendeur_B", "07 XX XX XX XX", "19/01/2026", "", "✅ Oui", "Mandat Signé", "Exclusif 3 mois"),
        (3, "20/01/2026", "PAP", "https://www.pap.fr/...", "M. Dupont_C", "06 XX XX XX XX", "21/01/2026", "", "❌ Non", "⚠️ À RÉGULARISER", "Demander capture"),
    ]
    
    for row_idx, data in enumerate(examples, 5):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_body(cell, alternate=(row_idx % 2 == 0))
            cell.alignment = ALIGN_CENTER if col_idx in [1, 2, 6, 7, 8, 9, 10] else ALIGN_LEFT
        
        # Formule expiration : Date Opt-in + 1095 jours (3 ans)
        ws.cell(row=row_idx, column=8, value=f"=IF(G{row_idx}<>\"\",G{row_idx}+1095,\"\")")
        ws.cell(row=row_idx, column=8).number_format = "DD/MM/YYYY"
    
    # === LIGNES VIDES POUR SAISIE (8-60) ===
    for row_idx in range(8, 61):
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            style_body(cell, alternate=(row_idx % 2 == 0))
            cell.alignment = ALIGN_CENTER if col_idx in [1, 2, 6, 7, 8, 9, 10] else ALIGN_LEFT
        
        # ID auto-incrémenté
        ws.cell(row=row_idx, column=1, value=f"=IF(B{row_idx}<>\"\",ROW()-4,\"\")")
        # Formule expiration
        ws.cell(row=row_idx, column=8, value=f"=IF(G{row_idx}<>\"\",G{row_idx}+1095,\"\")")
        ws.cell(row=row_idx, column=8).number_format = "DD/MM/YYYY"
    
    # === DATA VALIDATIONS ===
    # Plateforme
    dv_plateforme = DataValidation(
        type="list",
        formula1='"Leboncoin,SeLoger,PAP,Bien\'ici,Logic-Immo,Facebook,LinkedIn,Instagram,Autre"',
        allow_blank=True
    )
    dv_plateforme.prompt = "Choisir la plateforme"
    ws.add_data_validation(dv_plateforme)
    dv_plateforme.add('C5:C60')
    
    # Preuve stockée
    dv_preuve = DataValidation(
        type="list",
        formula1='"✅ Oui,❌ Non"',
        allow_blank=True
    )
    ws.add_data_validation(dv_preuve)
    dv_preuve.add('I5:I60')
    
    # Statut
    dv_statut = DataValidation(
        type="list",
        formula1='"En attente,RDV Planifié,Estimation faite,Négociation,Mandat Signé,Perdu,⚠️ À RÉGULARISER"',
        allow_blank=True
    )
    ws.add_data_validation(dv_statut)
    dv_statut.add('J5:J60')
    
    # === FILTRES ===
    ws.auto_filter.ref = "A4:K60"
    
    # === FIGER VOLETS ===
    ws.freeze_panes = "A5"
    
    # === LÉGENDE ===
    ws['A63'] = "📌 RAPPELS LÉGAUX :"
    ws['A63'].font = Font(name='Calibri', size=11, bold=True, color=BLEU_NUIT)
    
    ws.merge_cells('A64:K64')
    ws['A64'] = "• Le consentement doit être LIBRE, SPÉCIFIQUE, ÉCLAIRÉ et UNIVOQUE (RGPD)"
    ws['A64'].font = FONT_SUBTITLE
    
    ws.merge_cells('A65:K65')
    ws['A65'] = "• Preuve = Capture d'écran de la conversation OU export de la messagerie plateforme"
    ws['A65'].font = FONT_SUBTITLE
    
    ws.merge_cells('A66:K66')
    ws['A66'] = "• Sanction en cas de démarchage sans consentement : jusqu'à 75 000€ d'amende (personne physique)"
    ws['A66'].font = Font(name='Calibri', size=10, italic=True, color=ROUGE_ALERTE)
    
    wb.save(filepath)
    print(f"   ✅ {os.path.basename(filepath)}")


# =============================================================================
# 2. CALCULATEUR PRÊT RELAIS
# =============================================================================
def create_calculateur_pret_relais(filepath):
    """
    Calculateur coût du retard pour vendeur en prêt relais
    Objectif : Montrer l'impact financier de chaque mois perdu
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculateur Prêt Relais"
    
    # === TITRE ===
    ws.merge_cells('A1:E1')
    ws['A1'] = "💰 CALCULATEUR — LE COÛT DU RETARD (PRÊT RELAIS)"
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 35
    
    # === SOUS-TITRE ===
    ws.merge_cells('A2:E2')
    ws['A2'] = "Montrez à votre vendeur ce que lui coûte chaque mois de retard dans sa vente"
    ws['A2'].font = FONT_SUBTITLE
    ws['A2'].alignment = ALIGN_CENTER
    
    # === SECTION 1 : PARAMÈTRES ===
    ws['A4'] = "📝 PARAMÈTRES À RENSEIGNER"
    ws['A4'].font = Font(name='Calibri', size=12, bold=True, color=ACCENT_BLEU)
    ws.merge_cells('A4:C4')
    
    params = [
        (6, "Prix de vente estimé du bien", "B6", 350000, "€"),
        (7, "Capital restant dû (crédit en cours)", "B7", 180000, "€"),
        (8, "Taux annuel prêt relais (%)", "B8", 4.5, "%"),
        (9, "Charges mensuelles (copro + taxe foncière)", "B9", 350, "€/mois"),
        (10, "Crédit immobilier mensuel actuel", "B10", 1200, "€/mois"),
    ]
    
    for row, label, cell_ref, default, unit in params:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = FONT_BODY
        ws[f'A{row}'].alignment = ALIGN_LEFT
        
        ws[cell_ref] = default
        style_input(ws[cell_ref])
        
        ws[f'C{row}'] = unit
        ws[f'C{row}'].font = FONT_SUBTITLE
    
    # === SECTION 2 : CALCULS AUTOMATIQUES ===
    ws['A13'] = "📊 RÉSULTATS AUTOMATIQUES"
    ws['A13'].font = Font(name='Calibri', size=12, bold=True, color=VERT_SUCCES)
    ws.merge_cells('A13:C13')
    
    # Montant prêt relais
    ws['A15'] = "Montant du prêt relais"
    ws['A15'].font = FONT_BODY
    ws['B15'] = "=B6-B7"
    ws['B15'].font = FONT_ACCENT
    ws['B15'].number_format = '#,##0 €'
    ws['B15'].fill = FILL_RESULT
    ws['B15'].border = BORDER_THIN
    ws['C15'] = "(Prix - Capital dû)"
    ws['C15'].font = FONT_SUBTITLE
    
    # Intérêts mensuels prêt relais
    ws['A16'] = "Intérêts mensuels prêt relais"
    ws['A16'].font = FONT_BODY
    ws['B16'] = "=ROUND(B15*(B8/100)/12,2)"
    ws['B16'].font = FONT_ACCENT
    ws['B16'].number_format = '#,##0.00 €'
    ws['B16'].fill = FILL_RESULT
    ws['B16'].border = BORDER_THIN
    
    # Coût mensuel TOTAL
    ws['A18'] = "🔴 COÛT MENSUEL TOTAL DU RETARD"
    ws['A18'].font = Font(name='Calibri', size=11, bold=True, color=BLEU_NUIT)
    ws['B18'] = "=B16+B9+B10"
    ws['B18'].font = FONT_ALERT
    ws['B18'].number_format = '#,##0.00 €'
    ws['B18'].fill = FILL_ALERT
    ws['B18'].border = BORDER_ACCENT
    ws['C18'] = "/mois"
    ws['C18'].font = FONT_BODY
    
    # === SECTION 3 : SIMULATION RETARD ===
    ws['A21'] = "⏱️ SIMULATION : IMPACT DU DÉLAI DE VENTE"
    ws['A21'].font = Font(name='Calibri', size=12, bold=True, color=BLEU_NUIT)
    ws.merge_cells('A21:E21')
    
    # Headers simulation
    sim_headers = ["Délai supplémentaire", "Coût cumulé", "Équivalent", ""]
    for col_idx, header in enumerate(sim_headers, 1):
        cell = ws.cell(row=23, column=col_idx, value=header)
        style_header(cell)
    
    # Lignes simulation
    simulations = [
        ("+1 mois", "=B18*1", "1 week-end en famille"),
        ("+3 mois", "=B18*3", "Des vacances d'été"),
        ("+6 mois", "=B18*6", "Une voiture d'occasion"),
        ("+12 mois", "=B18*12", "Un an de crédit auto"),
    ]
    
    for row_idx, (delai, formule, equiv) in enumerate(simulations, 24):
        ws.cell(row=row_idx, column=1, value=delai)
        ws.cell(row=row_idx, column=1).font = FONT_BODY
        ws.cell(row=row_idx, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=1).border = BORDER_THIN
        
        ws.cell(row=row_idx, column=2, value=formule)
        ws.cell(row=row_idx, column=2).font = FONT_ALERT
        ws.cell(row=row_idx, column=2).number_format = '#,##0 €'
        ws.cell(row=row_idx, column=2).alignment = ALIGN_CENTER
        ws.cell(row=row_idx, column=2).border = BORDER_THIN
        ws.cell(row=row_idx, column=2).fill = FILL_ALERT if row_idx >= 26 else FILL_WARNING
        
        ws.cell(row=row_idx, column=3, value=equiv)
        ws.cell(row=row_idx, column=3).font = FONT_SUBTITLE
        ws.cell(row=row_idx, column=3).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=3).border = BORDER_THIN
    
    # === ARGUMENT CLÉ ===
    ws.merge_cells('A30:E30')
    ws['A30'] = "🎯 PHRASE CLÉ EN RENDEZ-VOUS :"
    ws['A30'].font = Font(name='Calibri', size=11, bold=True, color=BLEU_NUIT)
    
    ws.merge_cells('A31:E33')
    ws['A31'] = "« Chaque mois de retard vous coûte [B18] €. En 6 mois, c'est [=B18*6] € qui auraient pu rester dans votre poche. Mon objectif est de vous éviter ce surcoût en vendant efficacement, au bon prix. »"
    ws['A31'].font = Font(name='Calibri', size=10, italic=True, color="64748B")
    ws['A31'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A31'].fill = FILL_GRIS
    
    # === AJUSTER COLONNES ===
    adjust_columns(ws, {'A': 40, 'B': 20, 'C': 25, 'D': 15, 'E': 15})
    
    wb.save(filepath)
    print(f"   ✅ {os.path.basename(filepath)}")


# =============================================================================
# 3. SUIVI DPE — COEFFICIENT 2.3 → 1.9
# =============================================================================
def create_suivi_dpe(filepath):
    """
    Tableau de simulation DPE avec nouveau coefficient
    Formule : Nouvelle conso = Ancienne conso × (1.9 / 2.3)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi DPE"
    
    # === TITRE ===
    ws.merge_cells('A1:I1')
    ws['A1'] = "🔋 SIMULATEUR DPE — IMPACT COEFFICIENT 2.3 → 1.9 (1er janvier 2026)"
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 35
    
    # === AVERTISSEMENT ===
    ws.merge_cells('A3:I4')
    ws['A3'] = "⚠️ ATTENTION : Ce tableau fournit une ESTIMATION indicative. Seul un diagnostiqueur certifié peut délivrer un DPE officiel. L'amélioration réelle dépend des seuils, de la configuration du bien et de multiples paramètres."
    ws['A3'].font = Font(name='Calibri', size=9, color=ORANGE_WARNING)
    ws['A3'].alignment = Alignment(wrap_text=True, vertical='center')
    ws['A3'].fill = FILL_WARNING
    
    # === EXPLICATION ===
    ws.merge_cells('A6:I6')
    ws['A6'] = "📐 Formule appliquée : Nouvelle consommation = Consommation actuelle × (1.9 ÷ 2.3) ≈ réduction de ~17%"
    ws['A6'].font = FONT_SUBTITLE
    
    # === HEADERS ===
    headers = [
        ("A8", "Réf.", 8),
        ("B8", "Adresse / Description", 35),
        ("C8", "Surface\n(m²)", 10),
        ("D8", "Chauffage", 12),
        ("E8", "Conso Actuelle\n(kWh/m²/an)", 16),
        ("F8", "Lettre\nActuelle", 10),
        ("G8", "Conso Simulée\n(coef 1.9)", 16),
        ("H8", "Lettre\nEstimée", 10),
        ("I8", "Gain\nPotentiel", 12),
    ]
    
    for cell_ref, title, width in headers:
        ws[cell_ref] = title
        style_header(ws[cell_ref])
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width
    
    ws.row_dimensions[8].height = 40
    
    # === EXEMPLES ===
    examples = [
        ("B001", "Appt 3P Rue Victor Hugo", 68, "Électrique", 295, "E"),
        ("B002", "Maison 5P Avenue Foch", 145, "Électrique", 355, "F"),
        ("B003", "Studio Centre-Ville", 25, "Électrique", 245, "D"),
    ]
    
    for row_idx, (ref, adresse, surface, chauffage, conso, lettre) in enumerate(examples, 9):
        ws.cell(row=row_idx, column=1, value=ref)
        ws.cell(row=row_idx, column=2, value=adresse)
        ws.cell(row=row_idx, column=3, value=surface)
        ws.cell(row=row_idx, column=4, value=chauffage)
        ws.cell(row=row_idx, column=5, value=conso)
        ws.cell(row=row_idx, column=6, value=lettre)
        
        # Formule conso simulée
        ws.cell(row=row_idx, column=7, value=f"=ROUND(E{row_idx}*(1.9/2.3),0)")
        
        # Formule lettre estimée (seuils DPE 2021)
        ws.cell(row=row_idx, column=8, value=f'=IF(G{row_idx}="","",IF(G{row_idx}<=70,"A",IF(G{row_idx}<=110,"B",IF(G{row_idx}<=180,"C",IF(G{row_idx}<=250,"D",IF(G{row_idx}<=330,"E",IF(G{row_idx}<=420,"F","G")))))))')
        
        # Formule gain
        ws.cell(row=row_idx, column=9, value=f'=IF(OR(F{row_idx}="",H{row_idx}=""),"",IF(F{row_idx}=H{row_idx},"=",IF(CODE(H{row_idx})<CODE(F{row_idx}),"✅ +"&(CODE(F{row_idx})-CODE(H{row_idx}))&" classe","—")))')
        
        # Style
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            style_body(cell, alternate=(row_idx % 2 == 0))
            cell.alignment = ALIGN_CENTER
        
        ws.cell(row=row_idx, column=2).alignment = ALIGN_LEFT
        ws.cell(row=row_idx, column=7).font = FONT_ACCENT
        ws.cell(row=row_idx, column=7).fill = FILL_RESULT
        ws.cell(row=row_idx, column=8).font = FONT_ACCENT
        ws.cell(row=row_idx, column=9).font = Font(name='Calibri', size=10, bold=True, color=VERT_SUCCES)
    
    # === LIGNES VIDES POUR SAISIE ===
    for row_idx in range(12, 40):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            style_body(cell, alternate=(row_idx % 2 == 0))
            cell.alignment = ALIGN_CENTER
        
        ws.cell(row=row_idx, column=2).alignment = ALIGN_LEFT
        
        # Formules auto
        ws.cell(row=row_idx, column=7, value=f'=IF(E{row_idx}="","",ROUND(E{row_idx}*(1.9/2.3),0))')
        ws.cell(row=row_idx, column=7).font = FONT_ACCENT
        
        ws.cell(row=row_idx, column=8, value=f'=IF(G{row_idx}="","",IF(G{row_idx}<=70,"A",IF(G{row_idx}<=110,"B",IF(G{row_idx}<=180,"C",IF(G{row_idx}<=250,"D",IF(G{row_idx}<=330,"E",IF(G{row_idx}<=420,"F","G")))))))')
        
        ws.cell(row=row_idx, column=9, value=f'=IF(OR(F{row_idx}="",H{row_idx}=""),"",IF(F{row_idx}=H{row_idx},"=",IF(CODE(H{row_idx})<CODE(F{row_idx}),"✅ +"&(CODE(F{row_idx})-CODE(H{row_idx}))&" classe","—")))')
    
    # === DATA VALIDATIONS ===
    # Chauffage
    dv_chauffage = DataValidation(
        type="list",
        formula1='"Électrique,Gaz,Fioul,Bois,PAC,Autre"',
        allow_blank=True
    )
    ws.add_data_validation(dv_chauffage)
    dv_chauffage.add('D9:D40')
    
    # Lettre actuelle
    dv_lettre = DataValidation(
        type="list",
        formula1='"A,B,C,D,E,F,G"',
        allow_blank=True
    )
    ws.add_data_validation(dv_lettre)
    dv_lettre.add('F9:F40')
    
    # === FILTRES ===
    ws.auto_filter.ref = "A8:I40"
    
    # === FIGER VOLETS ===
    ws.freeze_panes = "A9"
    
    # === LÉGENDE SEUILS DPE ===
    ws['A43'] = "📊 SEUILS DPE (kWh EP/m²/an) — Méthode 2021"
    ws['A43'].font = Font(name='Calibri', size=11, bold=True, color=BLEU_NUIT)
    
    seuils = [
        ("A", "≤ 70", "22B14C"),
        ("B", "71-110", "50C878"),
        ("C", "111-180", "F7DC6F"),
        ("D", "181-250", "F5B041"),
        ("E", "251-330", "EB984E"),
        ("F", "331-420", "E74C3C"),
        ("G", "> 420", "943126"),
    ]
    
    for idx, (lettre, seuil, couleur) in enumerate(seuils):
        col = idx + 1
        cell_lettre = ws.cell(row=44, column=col, value=lettre)
        cell_lettre.font = Font(name='Calibri', size=11, bold=True, color=BLANC)
        cell_lettre.fill = PatternFill(start_color=couleur, end_color=couleur, fill_type='solid')
        cell_lettre.alignment = ALIGN_CENTER
        cell_lettre.border = BORDER_THIN
        
        cell_seuil = ws.cell(row=45, column=col, value=seuil)
        cell_seuil.font = Font(name='Calibri', size=8, color="64748B")
        cell_seuil.alignment = ALIGN_CENTER
    
    # === NOTE IMPORTANTE ===
    ws.merge_cells('A47:I48')
    ws['A47'] = "💡 NOTE : Cette simulation ne concerne QUE les logements chauffés à l'électricité. Les logements gaz/fioul ne sont pas impactés par ce changement de coefficient."
    ws['A47'].font = FONT_SUBTITLE
    ws['A47'].alignment = Alignment(wrap_text=True)
    
    wb.save(filepath)
    print(f"   ✅ {os.path.basename(filepath)}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    """Générer les 3 fichiers Excel"""
    output_dir = "/home/claude/kit-complet/04_Tableaux_Outils"
    os.makedirs(output_dir, exist_ok=True)
    
    print("\n" + "=" * 65)
    print("  🚀 KIT MANDATAIRE 2026 — GÉNÉRATION DES FICHIERS EXCEL")
    print("=" * 65)
    print(f"\n  📁 Dossier : {output_dir}\n")
    
    # Génération
    create_tracker_consentements(
        os.path.join(output_dir, "Tracker_Consentements_RGPD.xlsx")
    )
    create_calculateur_pret_relais(
        os.path.join(output_dir, "Calculateur_Pret_Relais.xlsx")
    )
    create_suivi_dpe(
        os.path.join(output_dir, "Suivi_DPE_Coefficient.xlsx")
    )
    
    print("\n" + "=" * 65)
    print("  ✅ TOUS LES FICHIERS EXCEL ONT ÉTÉ GÉNÉRÉS !")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
